#!/usr/bin/env python3
"""Export ACI-JP-Cardio human experiment results from the EMR SQLite DB.

The app writes the actual participant SOAP notes to `soap_notes` and timing /
AI interaction metadata to `experiment_attempts` and `experiment_events`.
This script makes a researcher-facing workbook without mutating the DB.
"""

from __future__ import annotations

import argparse
import csv
import json
import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Any


DEFAULT_DB_CANDIDATES = [
    Path("/tmp/emr-exp-run/ehr-demo.db"),
    Path("/home/junkanki/naka/emr/backend/ehr-demo.db"),
]

CASE_DOCS_MAP = {
    "C1": ("JC-AMI-A", 9, "急性心筋梗塞", "A"),
    "C2": ("JC-PE-B", 13, "急性肺塞栓症", "B"),
    "C3": ("JC-AD-A", 14, "大動脈解離", "A"),
    "C4": ("JC-AHF-B", 5, "急性心不全", "B"),
    "C5": ("JC-AMI-B", 3, "急性心筋梗塞", "B"),
    "C6": ("JC-PE-A", 12, "急性肺塞栓症", "A"),
    "C7": ("JC-AD-B", 15, "大動脈解離", "B"),
    "C8": ("JC-AHF-A", 20, "急性心不全", "A"),
}


def pick_default_db() -> Path:
    for path in DEFAULT_DB_CANDIDATES:
        if path.exists():
            return path
    return DEFAULT_DB_CANDIDATES[0]


def connect_readonly(db_path: Path) -> sqlite3.Connection:
    uri = f"file:{db_path}?mode=ro"
    conn = sqlite3.connect(uri, uri=True)
    conn.row_factory = sqlite3.Row
    return conn


def fetch_rows(conn: sqlite3.Connection, sql: str, params: tuple[Any, ...] = ()) -> list[dict[str, Any]]:
    return [dict(row) for row in conn.execute(sql, params)]


def as_text(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, (dict, list)):
        return json.dumps(v, ensure_ascii=False)
    return str(v)


def latest_results(conn: sqlite3.Connection) -> list[dict[str, Any]]:
    rows = fetch_rows(
        conn,
        """
        WITH latest_soap AS (
          SELECT *
          FROM (
            SELECT
              sn.*,
              ROW_NUMBER() OVER (
                PARTITION BY sn.encounter_id
                ORDER BY datetime(COALESCE(sn.updated_at, sn.created_at)) DESC, sn.id DESC
              ) AS rn
            FROM soap_notes sn
          )
          WHERE rn = 1
        ),
        event_counts AS (
          SELECT
            attempt_id,
            COUNT(*) AS event_count,
            SUM(CASE WHEN event_type = 'ai_draft_accepted' THEN 1 ELSE 0 END) AS ai_draft_accept_events,
            SUM(CASE WHEN event_type = 'ai_draft_edited' THEN 1 ELSE 0 END) AS ai_draft_edit_events,
            SUM(CASE WHEN event_type = 'ai_draft_rejected' THEN 1 ELSE 0 END) AS ai_draft_reject_events,
            SUM(CASE WHEN event_type = 'slm_autocomplete' THEN 1 ELSE 0 END) AS autocomplete_events,
            SUM(CASE WHEN event_type LIKE 'soap_draft%' THEN 1 ELSE 0 END) AS soap_draft_events
          FROM experiment_events
          GROUP BY attempt_id
        )
        SELECT
          ea.attempt_id,
          ea.subject_id,
          ea.sequence_order,
          ea.case_id,
          ea.source_case_id,
          ea.intervention,
          ea.status,
          ea.started_at,
          ea.ended_at,
          ea.duration_sec,
          ea.interruption_sec,
          ea.ai_wait_ms,
          ea.ai_candidate_count,
          ea.ai_accept_count,
          ea.ai_edit_count,
          ea.ai_reject_count,
          COALESCE(ec.event_count, 0) AS event_count,
          COALESCE(ec.ai_draft_accept_events, 0) AS ai_draft_accept_events,
          COALESCE(ec.ai_draft_edit_events, 0) AS ai_draft_edit_events,
          COALESCE(ec.ai_draft_reject_events, 0) AS ai_draft_reject_events,
          COALESCE(ec.autocomplete_events, 0) AS autocomplete_events,
          COALESCE(ec.soap_draft_events, 0) AS soap_draft_events,
          p.mrn,
          p.name AS patient_name,
          p.gender AS patient_gender,
          p.birth_date,
          e.encounter_date,
          e.encounter_type,
          e.department,
          e.chief_complaint,
          i.raw_text AS source_raw_text,
          i.medication_list AS source_medication_list,
          i.exam_findings AS source_exam_findings,
          i.lab_results AS source_lab_results,
          i.structured_data AS source_structured_data,
          ls.id AS soap_note_id,
          ls.author AS soap_author,
          ls.subjective,
          ls.objective,
          ls.assessment,
          ls.plan,
          ls.created_at AS soap_created_at,
          ls.updated_at AS soap_updated_at,
          CASE WHEN ls.id IS NULL THEN 0 ELSE 1 END AS has_saved_soap
        FROM experiment_attempts ea
        JOIN patients p ON p.id = ea.patient_id
        JOIN encounters e ON e.id = ea.encounter_id
        LEFT JOIN interview_notes i ON i.encounter_id = ea.encounter_id
        LEFT JOIN latest_soap ls ON ls.encounter_id = ea.encounter_id
        LEFT JOIN event_counts ec ON ec.attempt_id = ea.attempt_id
        ORDER BY ea.attempt_id
        """,
    )

    for row in rows:
        source_id, docs_no, disease, pattern = CASE_DOCS_MAP.get(
            row["case_id"],
            (row["source_case_id"], "", "", ""),
        )
        row["docs_no"] = docs_no
        row["disease"] = disease
        row["pattern"] = pattern
        row["source_case_id_check"] = source_id
        structured = parse_json(row.get("source_structured_data"))
        patient_info = structured.get("patient_info", {}) if isinstance(structured, dict) else {}
        row["source_patient_info_json"] = patient_info if isinstance(patient_info, dict) else {}
        if isinstance(patient_info, dict):
            row["source_patient_age"] = patient_info.get("age", "")
            row["source_patient_gender"] = patient_info.get("gender", "")
            row["source_secondary_complaints"] = join_list(patient_info.get("secondary_complaints"))
            row["source_comorbidities"] = join_list(patient_info.get("comorbidities"))
            row["source_current_medications"] = join_list(patient_info.get("current_medications"))
            row["source_allergies"] = join_list(patient_info.get("allergies"))
            row["source_family_history"] = join_list(patient_info.get("family_history"))
            row["source_social_history"] = patient_info.get("social_history", "")
            row["source_reception_vitals"] = patient_info.get("reception_vitals", {})
        row["soap_full_text"] = "\n\n".join(
            part
            for part in [
                f"S:\n{row.get('subjective') or ''}".strip(),
                f"O:\n{row.get('objective') or ''}".strip(),
                f"A:\n{row.get('assessment') or ''}".strip(),
                f"P:\n{row.get('plan') or ''}".strip(),
            ]
            if part
        )
    return rows


def parse_json(value: Any) -> Any:
    if not value:
        return None
    if isinstance(value, (dict, list)):
        return value
    try:
        return json.loads(str(value))
    except json.JSONDecodeError:
        return None


def join_list(value: Any) -> str:
    if isinstance(value, list):
        return " / ".join(str(v) for v in value)
    return as_text(value)


def all_soap_notes(conn: sqlite3.Connection) -> list[dict[str, Any]]:
    return fetch_rows(
        conn,
        """
        SELECT
          ea.attempt_id,
          ea.subject_id,
          ea.sequence_order,
          ea.case_id,
          ea.intervention,
          sn.id AS soap_note_id,
          sn.author,
          sn.subjective,
          sn.objective,
          sn.assessment,
          sn.plan,
          sn.created_at,
          sn.updated_at
        FROM experiment_attempts ea
        JOIN soap_notes sn ON sn.encounter_id = ea.encounter_id
        ORDER BY ea.attempt_id, datetime(COALESCE(sn.updated_at, sn.created_at)), sn.id
        """,
    )


def events(conn: sqlite3.Connection) -> list[dict[str, Any]]:
    return fetch_rows(
        conn,
        """
        SELECT
          ev.id,
          ev.attempt_id,
          ea.subject_id,
          ea.case_id,
          ea.intervention,
          ev.event_type,
          ev.payload_json,
          ev.created_at
        FROM experiment_events ev
        LEFT JOIN experiment_attempts ea ON ea.attempt_id = ev.attempt_id
        ORDER BY ev.created_at, ev.id
        """,
    )


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if rows:
        fields = list(rows[0].keys())
    else:
        fields = []
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            writer.writerow({k: as_text(v) for k, v in row.items()})


def write_xlsx(path: Path, sheets: dict[str, list[dict[str, Any]]]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise SystemExit("openpyxl is required for .xlsx export. Use --csv-dir instead.") from exc

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    for name, rows in sheets.items():
        ws = wb.create_sheet(title=name[:31])
        fields = list(rows[0].keys()) if rows else []
        if fields:
            ws.append(fields)
            for cell in ws[1]:
                cell.font = Font(bold=True)
            for row in rows:
                ws.append([as_text(row.get(field)) for field in fields])
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            for idx, field in enumerate(fields, start=1):
                width = 14
                if field in {
                    "subjective",
                    "objective",
                    "assessment",
                    "plan",
                    "soap_full_text",
                    "source_raw_text",
                    "source_patient_info_json",
                    "source_social_history",
                }:
                    width = 48
                elif field in {"chief_complaint", "payload_json"}:
                    width = 36
                ws.column_dimensions[get_column_letter(idx)].width = width
        else:
            ws.append(["no rows"])

    wb.save(path)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--db", type=Path, default=pick_default_db(), help="SQLite DB path")
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("/data2/junkanki/naka/exports")
        / f"experiment_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        help="Output .xlsx path",
    )
    parser.add_argument("--csv-dir", type=Path, help="Optional directory for CSV copies")
    args = parser.parse_args()

    with connect_readonly(args.db) as conn:
        sheets = {
            "results": latest_results(conn),
            "all_soap_notes": all_soap_notes(conn),
            "events": events(conn),
        }

    if args.output.suffix.lower() == ".xlsx":
        write_xlsx(args.output, sheets)
        print(args.output)
    else:
        raise SystemExit("--output must end with .xlsx")

    if args.csv_dir:
        for name, rows in sheets.items():
            write_csv(args.csv_dir / f"{name}.csv", rows)
        print(args.csv_dir)


if __name__ == "__main__":
    main()
