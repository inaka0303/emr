#!/usr/bin/env python3
"""Merge manual LLM-judge JSONL results into the experiment score workbook."""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

import export_experiment_results as export_results


DEFAULT_SCORES_XLSX = Path("/home/junkanki/naka/emr/exports/experiment_scores_current.xlsx")
DEFAULT_JUDGE_RESULTS = Path("/home/junkanki/naka/emr/exports/judge_results_current.jsonl")
DEFAULT_OUTPUT = Path("/home/junkanki/naka/emr/exports/experiment_scores_with_judge_current.xlsx")

JUDGE_AXES = [
    "medical",
    "completeness",
    "naturalness",
    "hallucination_absence",
    "format",
]


def average(values: list[float | None]) -> float | None:
    numeric = [v for v in values if isinstance(v, (int, float))]
    return round(sum(numeric) / len(numeric), 4) if numeric else None


def to_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text or text.lower() in {"none", "null", "nan"}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def as_list_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, list):
        return " / ".join(str(v) for v in value)
    return str(value)


def read_xlsx(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit("openpyxl is required to read score workbooks") from exc

    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise SystemExit(f"sheet not found: {sheet_name} in {path}")
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(v) if v is not None else "" for v in rows[0]]
    out: list[dict[str, Any]] = []
    for values in rows[1:]:
        if not any(v is not None and str(v).strip() for v in values):
            continue
        out.append({headers[idx]: values[idx] if idx < len(values) else "" for idx in range(len(headers))})
    return out


def parse_json_objects(path: Path) -> list[dict[str, Any]]:
    text = path.read_text(encoding="utf-8").strip()
    if not text:
        return []

    try:
        parsed = json.loads(text)
        if isinstance(parsed, list):
            return [obj for obj in parsed if isinstance(obj, dict)]
        if isinstance(parsed, dict):
            return [parsed]
    except json.JSONDecodeError:
        pass

    rows: list[dict[str, Any]] = []
    for line_no, line in enumerate(text.splitlines(), start=1):
        line = line.strip()
        if not line or line.startswith("```"):
            continue
        try:
            obj = json.loads(line)
        except json.JSONDecodeError as exc:
            raise SystemExit(f"invalid JSON on line {line_no}: {exc}") from exc
        if not isinstance(obj, dict):
            raise SystemExit(f"line {line_no} is not a JSON object")
        rows.append(obj)
    return rows


def normalize_judge(row: dict[str, Any]) -> dict[str, Any]:
    attempt_id = str(row.get("attempt_id") or "").strip()
    if not attempt_id:
        raise SystemExit(f"judge row missing attempt_id: {row}")

    out: dict[str, Any] = {
        "attempt_id": attempt_id,
        "judge_model": row.get("judge_model") or row.get("model") or "",
        "judge_rubric_version": row.get("rubric_version") or row.get("judge_rubric_version") or "",
        "judge_major_omissions": as_list_text(row.get("major_omissions")),
        "judge_unsafe_or_hallucinated_items": as_list_text(row.get("unsafe_or_hallucinated_items")),
        "judge_comment": row.get("comment") or "",
    }

    axis_values: list[float] = []
    for axis in JUDGE_AXES:
        value = to_float(row.get(axis))
        if value is None or not 1 <= value <= 5:
            raise SystemExit(f"judge score for {attempt_id}.{axis} must be 1-5: {row.get(axis)!r}")
        out[f"judge_{axis}"] = int(value) if float(value).is_integer() else value
        axis_values.append(float(value))

    out["judge_mean_5"] = average(axis_values)
    out["judge_score_norm"] = round(float(out["judge_mean_5"]) / 5.0, 4)
    return out


def load_judges(path: Path) -> dict[str, dict[str, Any]]:
    judges: dict[str, dict[str, Any]] = {}
    for row in parse_json_objects(path):
        normalized = normalize_judge(row)
        judges[normalized["attempt_id"]] = normalized
    return judges


def merge_scores(scores: list[dict[str, Any]], judges: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    merged: list[dict[str, Any]] = []
    for row in scores:
        out = dict(row)
        attempt_id = str(row.get("attempt_id") or "")
        judge = judges.get(attempt_id)
        if judge:
            out["judge_status"] = "judged"
            out.update(judge)
            components = [
                to_float(out.get("rouge_l")),
                to_float(out.get("bertscore_f1")),
                to_float(out.get("drug_f1")),
                to_float(out.get("diagnosis_f1")),
                to_float(out.get("vitals_match_rate")),
                to_float(out.get("judge_score_norm")),
            ]
            out["composite_with_judge"] = average(components)
        else:
            out["judge_status"] = "missing_judge"
            for axis in JUDGE_AXES:
                out[f"judge_{axis}"] = ""
            out["judge_mean_5"] = ""
            out["judge_score_norm"] = ""
            out["judge_major_omissions"] = ""
            out["judge_unsafe_or_hallucinated_items"] = ""
            out["judge_comment"] = ""
            out["composite_with_judge"] = ""
        merged.append(out)
    return merged


def summarize(rows: list[dict[str, Any]], group_key: str) -> list[dict[str, Any]]:
    groups: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        groups.setdefault(str(row.get(group_key) or ""), []).append(row)

    out: list[dict[str, Any]] = []
    for key, members in sorted(groups.items()):
        scored = [m for m in members if m.get("score_status") == "scored"]
        judged = [m for m in scored if m.get("judge_status") == "judged"]
        out.append(
            {
                group_key: key,
                "attempts": len(members),
                "scored": len(scored),
                "judged": len(judged),
                "duration_sec": average([to_float(m.get("duration_sec")) for m in scored]),
                "composite_score": average([to_float(m.get("composite_score")) for m in scored]),
                "composite_with_judge": average([to_float(m.get("composite_with_judge")) for m in judged]),
                "judge_score_norm": average([to_float(m.get("judge_score_norm")) for m in judged]),
                "judge_medical": average([to_float(m.get("judge_medical")) for m in judged]),
                "judge_completeness": average([to_float(m.get("judge_completeness")) for m in judged]),
                "judge_naturalness": average([to_float(m.get("judge_naturalness")) for m in judged]),
                "judge_hallucination_absence": average(
                    [to_float(m.get("judge_hallucination_absence")) for m in judged]
                ),
                "judge_format": average([to_float(m.get("judge_format")) for m in judged]),
                "rouge_l": average([to_float(m.get("rouge_l")) for m in scored]),
                "drug_f1": average([to_float(m.get("drug_f1")) for m in scored]),
                "diagnosis_f1": average([to_float(m.get("diagnosis_f1")) for m in scored]),
                "vitals_match_rate": average([to_float(m.get("vitals_match_rate")) for m in scored]),
            }
        )
    return out


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--scores-xlsx", type=Path, default=DEFAULT_SCORES_XLSX, help="Current score workbook")
    parser.add_argument("--judge-results", type=Path, default=DEFAULT_JUDGE_RESULTS, help="Manual judge JSONL")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT, help="Merged output workbook")
    args = parser.parse_args()

    if not args.scores_xlsx.exists():
        raise SystemExit(f"score workbook not found: {args.scores_xlsx}")
    if not args.judge_results.exists():
        raise SystemExit(f"judge results JSONL not found: {args.judge_results}")

    scores = read_xlsx(args.scores_xlsx, "scores")
    judges = load_judges(args.judge_results)
    merged = merge_scores(scores, judges)

    export_results.write_xlsx(
        args.output,
        {
            "scores_with_judge": merged,
            "summary_by_intervention": summarize(merged, "intervention"),
            "summary_by_case": summarize(merged, "case_id"),
            "summary_by_subject": summarize(merged, "subject_id"),
            "judge_raw": list(judges.values()),
        },
    )
    print(args.output)
    print(f"merged {len(judges)} judge rows into {len(scores)} score rows")


if __name__ == "__main__":
    main()
